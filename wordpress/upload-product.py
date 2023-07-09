
from selenium import webdriver
from openpyxl import load_workbook
from time import sleep
from selenium.webdriver.common.keys import Keys
import random
import pandas as pd

# Open the Excel file
import openpyxl
excel_file_path = '/Users/khangnt/product.xlsx'  # Replace with the actual path to your Excel file
workbook = openpyxl.load_workbook(excel_file_path)


# Select the desired sheet (assuming the first sheet here)
sheet = workbook.active

options = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values.notifications" : 2}
options.add_experimental_option("prefs",prefs)
options.add_argument("start-maximized")
browser = webdriver.Chrome('/Users/khangnt/Downloads/chromedriver_mac_arm64/chromedriver', options=options)

def wait():
    return sleep(random.randint(5,8))


browser.get("")
wait()

id = browser.find_element_by_xpath('//*[@id="user_login"]')
id.send_keys("")
pwd = browser.find_element_by_xpath('//*[@id="user_pass"]')
pwd.send_keys("")
pwd.send_keys(Keys.ENTER)

sleep(3)

button = browser.find_element_by_xpath('//*[@id="menu-posts-product"]/a/div[3]')
# Click the button
button.click()

sleep(3)


# Click button Sản phẩm
button = browser.find_element_by_xpath('//*[@id="menu-posts-product"]/a/div[3]')
# Click the button
button.click()

file_path = '/Users/khangnt/product.xls'
df = pd.read_excel(file_path)


for row in sheet.iter_rows(values_only=True):

    browser.get("http://hmmobilehanoi.com/wp-admin/edit.php?post_type=product")

    # Click button Thêm mới
    button = browser.find_element_by_xpath('//*[@id="wpbody-content"]/div[6]/a[1]')
    # Click the button
    button.click()

    sleep(1)

    id = browser.find_element_by_xpath('//*[@id="title"]')
    id.send_keys(row[0])

    print(row[0])

    sleep(1)

    id = browser.find_element_by_xpath('//*[@id="_regular_price"]')
    id.send_keys(int(row[1])*1000)

    print(int(row[1])*1000)

    browser.execute_script("window.scrollTo(0, 0);")

    sleep(1)

    # Check the value of row[2]
    if row[2] == "samsung":
        # Samsung
        button = browser.find_element_by_xpath('// *[ @ id = "in-product_cat-2056"]')
        # Click the button
        button.click()
    elif row[2] == 'Oppo':
        # Oppo
        button = browser.find_element_by_xpath('// *[ @ id = "in-product_cat-2057"]')
        # Click the button
        button.click()
    elif row[2] == 'Realme':
        # Realme
        button = browser.find_element_by_xpath('// *[ @ id = "in-product_cat-2058"]')
        # Click the button
        button.click()
    elif row[2] == 'phu kien':
        # Phu kien
        button = browser.find_element_by_xpath('// *[ @ id = "in-product_cat-2054"]')
        # Click the button
        button.click()

    print(row[2])


    sleep(1)

    button = browser.find_element_by_xpath('//*[@id="publish"]')
    # Click the button
    button.click()

    sleep(1)

    print("Done")




browser.quit()